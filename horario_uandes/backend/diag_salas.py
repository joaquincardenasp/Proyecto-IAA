"""
diag_salas.py — Diagnóstico de la restricción de sala especial (RD4).

Ejecutar desde backend/:
    python diag_salas.py

Responde:
  1. ¿Qué hay en datos.capacidad_por_sala?
  2. ¿Qué valores de sala_especial tienen los cursos?
  3. ¿Hay coincidencia entre ambos?
  4. ¿Qué secciones de Plan Común tienen sala especial y cuántas son?
  5. ¿Cuáles causarían INFEASIBLE con capacity=1?
"""
import sys
from pathlib import Path
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

from app.core.parser import cargar_datos

INPUTS_DIR = Path(__file__).parent / "inputs"

print("Cargando datos…")
datos = cargar_datos(INPUTS_DIR)

# ── 1. capacidad_por_sala ──────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("1. datos.capacidad_por_sala")
print("=" * 60)
if datos.capacidad_por_sala:
    for nombre, cap in sorted(datos.capacidad_por_sala.items()):
        print(f"  '{nombre}': {cap} sala(s)")
else:
    print("  *** VACÍO — SALAS_ESPECIALES_ING.xlsx no encontrado o TIPO no leído ***")

# ── 2. sala_especial en cursos ────────────────────────────────────────────────
print("\n" + "=" * 60)
print("2. Valores de Curso.sala_especial en los datos")
print("=" * 60)
salas_usadas: dict[str, list[str]] = defaultdict(list)
for codigo, curso in datos.cursos.items():
    if curso.sala_especial:
        salas_usadas[curso.sala_especial].append(codigo)

if salas_usadas:
    for sala, codigos in sorted(salas_usadas.items()):
        cap = datos.capacidad_por_sala.get(sala)
        match = "✓ match" if cap is not None else "✗ SIN MATCH en capacidad_por_sala"
        print(f"  '{sala}' → {len(codigos)} cursos — capacidad: {cap if cap else 'desconocida'} {match}")
        if len(codigos) <= 6:
            print(f"     Cursos: {', '.join(codigos)}")
else:
    print("  No hay cursos con sala_especial.")

# ── 3. Secciones Plan Común con sala especial ─────────────────────────────────
print("\n" + "=" * 60)
print("3. Secciones de Plan Común con sala especial")
print("=" * 60)
codigos_pc = {
    c.codigo for c in datos.cursos.values()
    if "Plan Común" in c.semestres_por_carrera
}
sec_by_id = {s.id: s for s in datos.secciones}

por_sala_pc: dict[str, list] = defaultdict(list)
for s in datos.secciones:
    if s.codigo_curso not in codigos_pc:
        continue
    curso = datos.cursos.get(s.codigo_curso)
    if curso and curso.sala_especial:
        por_sala_pc[curso.sala_especial].append(s)

if por_sala_pc:
    for sala, secs in sorted(por_sala_pc.items()):
        cap_real = datos.capacidad_por_sala.get(sala, None)
        n_secs = len(secs)

        # Contar secciones del mismo curso-componente (RC las forzará al mismo bloque)
        grupos: dict[tuple, list] = defaultdict(list)
        for s in secs:
            grupos[(s.codigo_curso, s.componente.value)].append(s.id)

        max_concurrentes = max(len(ids) for ids in grupos.values())
        print(f"\n  Sala: '{sala}'")
        print(f"    Capacidad en datos: {cap_real if cap_real else 'desconocida (default=1)'}")
        print(f"    Total secciones: {n_secs}")
        print(f"    Max secciones del mismo grupo (RC → mismo bloque): {max_concurrentes}")

        if cap_real is None:
            if max_concurrentes > 1:
                print(f"    *** PROBLEMA: {max_concurrentes} secciones del mismo curso forzadas")
                print(f"        al mismo bloque por RC, pero capacidad desconocida (default=1)")
                print(f"        → RD4 las pondrá en conflicto → INFEASIBLE ***")
        elif max_concurrentes > cap_real:
            print(f"    *** INFEASIBLE ESPERADO: {max_concurrentes} secciones > {cap_real} salas ***")
        else:
            print(f"    OK: {max_concurrentes} secciones concurrentes ≤ {cap_real} salas")

        for (codigo, comp), ids in sorted(grupos.items()):
            print(f"      {codigo} {comp}: {len(ids)} secciones ({', '.join(s[-4:] for s in ids[:4])}...)"
                  if len(ids) > 4 else f"      {codigo} {comp}: {len(ids)} secciones")
else:
    print("  Ninguna sección de Plan Común tiene sala especial.")

# ── 4. Verificar si SALAS_ESPECIALES existe ───────────────────────────────────
print("\n" + "=" * 60)
print("4. Archivos de salas en inputs/")
print("=" * 60)
salas_file = INPUTS_DIR / "SALAS_ESPECIALES_ING.xlsx"
if salas_file.exists():
    import openpyxl
    wb = openpyxl.load_workbook(salas_file, read_only=True)
    print(f"  Archivo: {salas_file.name}")
    print(f"  Hojas: {wb.sheetnames}")
    # Leer hoja SALAS ESPECIALES
    if "SALAS ESPECIALES" in wb.sheetnames:
        ws = wb["SALAS ESPECIALES"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  Columnas: {headers}")
        tipos_vistos: dict[str, int] = defaultdict(int)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(v is not None for v in row):
                tipo = str(row[headers.index("TIPO")] if "TIPO" in headers else "").strip()
                if tipo:
                    tipos_vistos[tipo] += 1
        print("\n  Tipos en hoja SALAS ESPECIALES (TIPO → count):")
        for tipo, cnt in sorted(tipos_vistos.items()):
            match = "✓" if tipo in salas_usadas else "✗ no usado por ningún curso"
            print(f"    '{tipo}': {cnt} sala(s) física(s) — {match}")
    if "BBDD" in wb.sheetnames:
        ws_bbdd = wb["BBDD"]
        headers_bbdd = [c.value for c in next(ws_bbdd.iter_rows(min_row=1, max_row=1))]
        print(f"\n  Columnas hoja BBDD: {headers_bbdd}")
        print("  Primeras 5 filas de BBDD:")
        for i, row in enumerate(ws_bbdd.iter_rows(min_row=2, max_row=6, values_only=True)):
            print(f"    {row}")
else:
    print(f"  *** {salas_file.name} NO encontrado en inputs/ ***")

print("\n" + "=" * 60)
print("DIAGNÓSTICO COMPLETO")
print("=" * 60)
