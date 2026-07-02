"""
test_solver_step9.py — Paso 9: API FastAPI.

Verifica:
  - GET  /api/health           → 200 {"status": "ok"}
  - GET  /api/status           → 200, status=idle antes de solve
  - POST /api/solve            → 202 (inicia solver en background)
  - Polling /api/status        → termina en "ready" o "error"
  - GET  /api/results          → 200, retorna secciones y métricas válidas
  - GET  /api/export           → 200, bytes de un .xlsx válido
  - POST /api/solve de nuevo   → 202 (puede reiniciarse)
  - POST /api/solve concurrente → 409 (no puede haber dos ejecuciones)

Ejecutar desde backend/:
    python tests/test_solver_step9.py
"""
import sys
import time
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.api.routes import _state

client = TestClient(app)

TIMEOUT_SEC = 600   # CP-SAT + GA pueden tardar varios minutos


def check(cond: bool, msg: str) -> None:
    estado = "✓" if cond else "✗ FALLO"
    print(f"  [{estado}] {msg}")
    if not cond:
        raise AssertionError(f"FALLO: {msg}")


def wait_ready(timeout: int = TIMEOUT_SEC) -> str:
    """Hace polling a /api/status hasta ready/error o timeout."""
    deadline = time.time() + timeout
    last_progress = ""
    while time.time() < deadline:
        r = client.get("/api/status")
        data = r.json()
        status = data["status"]
        progress = data.get("progress", "")
        if progress != last_progress:
            print(f"    status={status}  progress={progress}")
            last_progress = progress
        if status in ("ready", "error"):
            return status
        time.sleep(2)
    return "timeout"


def test_step9():
    print("\n--- test_step9 (API FastAPI) ---")

    # 1. Health
    r = client.get("/api/health")
    check(r.status_code == 200, "GET /api/health → 200")
    check(r.json()["status"] == "ok", "health.status == 'ok'")

    # 2. Status inicial
    r = client.get("/api/status")
    check(r.status_code == 200, "GET /api/status → 200")
    check(r.json()["status"] == "idle", "status inicial == 'idle'")

    # 3. Results antes de solve → 404
    r = client.get("/api/results")
    check(r.status_code == 404, "GET /api/results sin solve → 404")

    # 4. Export antes de solve → 404
    r = client.get("/api/export")
    check(r.status_code == 404, "GET /api/export sin solve → 404")

    # 5. Intento de solve concurrente → 409 (simulamos estado "running")
    # TestClient ejecuta background tasks sincrónicamente, así que probamos el 409
    # directamente con un override de estado antes de lanzar el solve real.
    _state["status"] = "running"
    r2 = client.post("/api/solve", json={"n_generaciones": 50, "pop_size": 20})
    check(r2.status_code == 409, "POST /api/solve con estado 'running' → 409")
    _state["status"] = "idle"  # restaurar

    # 6. Lanzar solver real (pocos parámetros para que sea rápido)
    print("\n  Lanzando solver (n_generaciones=50, pop_size=20)…")
    payload = {"n_generaciones": 50, "pop_size": 20, "tiempo_limite_cpsat": 30.0}
    r = client.post("/api/solve", json=payload)
    check(r.status_code == 202, "POST /api/solve → 202")

    # Con TestClient, los background tasks se ejecutan sincrónicamente,
    # así que al retornar el solve ya ha terminado — no hace falta polling.
    r_status = client.get("/api/status")
    final_status = r_status.json()["status"]
    print(f"  status tras solve: {final_status}  progress: {r_status.json().get('progress','')}")
    check(final_status == "ready", f"Solver terminó con status='ready' (got={final_status})")

    if final_status != "ready":
        print(f"  Error: {r_status.json()['error']}")
        return

    # 8. Results
    r = client.get("/api/results")
    check(r.status_code == 200, "GET /api/results → 200")
    data = r.json()
    check("metricas" in data, "results tiene 'metricas'")
    check("secciones" in data, "results tiene 'secciones'")

    metricas = data["metricas"]
    check(metricas["n_secciones"] > 0, f"n_secciones > 0 ({metricas['n_secciones']})")
    check(metricas["fitness_ga"] >= 0, "fitness_ga >= 0")
    check(metricas["estado_cpsat"] in ("OPTIMAL", "FEASIBLE"),
          f"estado_cpsat válido ({metricas['estado_cpsat']})")

    secciones = data["secciones"]
    check(len(secciones) == metricas["n_secciones"],
          f"len(secciones)==n_secciones ({len(secciones)})")

    # Nuevo contrato del asistente: estado global + diagnóstico si es parcial
    estado = data.get("estado")
    check(estado in ("FACTIBLE", "PARCIAL", "INFEASIBLE"),
          f"results tiene estado válido ({estado})")
    if estado == "PARCIAL":
        diag = data.get("diagnostico")
        check(diag is not None and len(diag.get("unidades", [])) > 0,
              "un resultado PARCIAL incluye diagnóstico con unidades bloqueadas")

    # Validar estructura de una sección
    s0 = secciones[0]
    for campo in ("id", "codigo", "titulo", "seccion", "tipo", "profesor", "bloques"):
        check(campo in s0, f"seccion[0] tiene campo '{campo}'")
    check(len(s0["bloques"]) > 0, "seccion[0] tiene al menos un bloque")
    b0 = s0["bloques"][0]
    for campo in ("dia", "hora_inicio", "hora_fin", "tipo_bloque"):
        check(campo in b0, f"bloque[0] tiene campo '{campo}'")

    print(f"\n  Secciones: {len(secciones)}   "
          f"fitness GA: {metricas['fitness_ga']:.0f}   "
          f"mejora: {metricas['mejora_pct']:.1f}%")

    # 9. Export
    r = client.get("/api/export")
    check(r.status_code == 200, "GET /api/export → 200")
    check(r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "content-type es xlsx")
    check(len(r.content) > 1000, f"Excel tiene tamaño razonable ({len(r.content):,} bytes)")

    # Validar que el xlsx es abrible
    import io
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    check("Horario" in wb.sheetnames, "Excel contiene hoja 'Horario'")
    check("Métricas" in wb.sheetnames, "Excel contiene hoja 'Métricas'")
    print(f"  Excel válido, hojas: {wb.sheetnames}")

    # 10. Re-lanzar solve desde "ready" → 202 (se puede relanzar)
    r = client.post("/api/solve", json={"n_generaciones": 10, "pop_size": 10})
    check(r.status_code == 202, "POST /api/solve desde 'ready' → 202")

    print("\n  → step9 OK")


def main():
    fallidos = []
    try:
        test_step9()
    except AssertionError as e:
        fallidos.append(str(e))
    except Exception as e:
        fallidos.append(f"ERROR inesperado: {e}")
        import traceback
        traceback.print_exc()

    print()
    if fallidos:
        print(f"RESULTADO: {len(fallidos)} test(s) FALLARON:")
        for msg in fallidos:
            print(f"  ✗ {msg}")
        sys.exit(1)
    else:
        print("RESULTADO: PASO 9 VALIDADO ✓")


if __name__ == "__main__":
    main()
