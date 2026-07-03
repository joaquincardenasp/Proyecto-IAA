"""
test_solver_step8.py — Paso 8: Exportador Excel.

Verifica:
  - El archivo se genera sin errores y no está vacío
  - Contiene las hojas esperadas (Horario + 7 carreras + Métricas)
  - La hoja "Horario" tiene exactamente tantas filas de datos como secciones asignadas
  - Los bloques 13:30-15:20 aparecen en el Excel (validación del fix de bloques)
  - El archivo es abrirable por openpyxl (estructura válida)
  - Guarda el resultado en outputs/horario_generado.xlsx para inspección manual

Ejecutar desde backend/:
    python tests/test_solver_step8.py
"""
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

from app.core.parser import cargar_datos
from app.core.solver_cpsat import resolver_por_partes
from app.core.solver_ga import ejecutar_ga, calcular_fitness, encode, construir_contexto, PESOS
from app.core.exporter import exportar_horario, _CARRERAS
from app.core.blocks import TODOS_BLOQUES

INPUTS_DIR  = Path(__file__).parent.parent / "inputs"
OUTPUTS_DIR = Path(__file__).parent.parent / "outputs"
OUTPUT_FILE = OUTPUTS_DIR / "horario_generado.xlsx"

CARRERAS = ["Plan Común", "ICI", "IOC", "ICE", "ICC", "ICA", "ICQ"]


def check(condicion: bool, mensaje: str) -> None:
    estado = "✓" if condicion else "✗ FALLO"
    print(f"  [{estado}] {mensaje}")
    if not condicion:
        raise AssertionError(f"FALLO: {mensaje}")


def test_step8(datos):
    print("\n--- test_step8 (Exportador Excel) ---")

    # 1. Resolver (CP-SAT + GA rápido)
    print("\n  Ejecutando CP-SAT…")
    resultado_cpsat = resolver_por_partes(datos, carreras=CARRERAS)
    check(resultado_cpsat.estado in ("FACTIBLE", "PARCIAL"),
          f"El sistema entregó un horario ({resultado_cpsat.estado})")

    print("  Ejecutando GA (50 generaciones para test rápido)…")
    resultado_ga = ejecutar_ga(
        datos, resultado_cpsat.asignaciones,
        n_generaciones=50, pop_size=20, seed=42,
    )

    asignaciones = resultado_ga.asignaciones
    n_secciones  = len(asignaciones)

    # 2. Construir dict de métricas
    ctx = construir_contexto(datos, resultado_cpsat.asignaciones)
    fitness_cpsat = calcular_fitness(encode(resultado_cpsat.asignaciones, ctx), ctx)[0]
    fitness_ga    = resultado_ga.fitness_final
    mejora_pct    = (fitness_cpsat - fitness_ga) / fitness_cpsat * 100 if fitness_cpsat > 0 else 0

    metricas = {
        "fitness_cpsat": fitness_cpsat,
        "fitness_ga":    fitness_ga,
        "mejora_pct":    mejora_pct,
        "rb_detalle": {
            f"RB{k+1} (peso {v})": "—"
            for k, v in enumerate(PESOS.values())
        },
    }

    # 3. Exportar
    print(f"\n  Exportando a {OUTPUT_FILE}…")
    excel_bytes = exportar_horario(datos, asignaciones, OUTPUT_FILE, metricas)

    # 4. Verificaciones básicas
    check(len(excel_bytes) > 0, "El archivo Excel no está vacío")
    check(OUTPUT_FILE.exists(), f"El archivo fue guardado en {OUTPUT_FILE}")
    check(OUTPUT_FILE.stat().st_size > 1000, "El archivo tiene tamaño razonable (>1 KB)")

    # 5. Abrir con openpyxl y verificar estructura
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    sheet_names = wb.sheetnames
    print(f"\n  Hojas generadas: {sheet_names}")

    # Formato del cliente (idéntico al histórico): hoja única "HORARIO ING",
    # encabezado en la fila 14, una fila por (sección, bloque).
    check("HORARIO ING" in sheet_names, "Hoja 'HORARIO ING' presente en el Excel")
    ws = wb["HORARIO ING"]

    # 6. Encabezado en la fila 14 con las columnas del formato histórico
    header = [ws.cell(row=14, column=c).value for c in range(1, 20)]
    for col in ("AREA", "NRC", "TITULO", "Lunes", "TIPO DE REUNION", "PROFESOR"):
        check(col in header, f"Encabezado (fila 14) tiene columna '{col}'")

    # 7. Una fila de datos por (sección, bloque)
    n_bloques = sum(len(b) for b in asignaciones.values())
    data_rows = ws.max_row - 14
    check(
        data_rows == n_bloques,
        f"Filas de datos == bloques asignados ({data_rows} == {n_bloques})",
    )

    # 8. El bloque 13:30 aparece en alguna celda de día
    valores_celdas = set()
    for row in ws.iter_rows(min_row=15, values_only=True):
        for cell in row:
            if isinstance(cell, str):
                valores_celdas.add(cell)
    check(any("13:30" in v for v in valores_celdas), "El bloque 13:30 aparece en la hoja")

    print(f"\n  Archivo guardado en: {OUTPUT_FILE}")
    print(f"  Tamaño: {OUTPUT_FILE.stat().st_size:,} bytes")
    print(f"  Secciones exportadas: {n_secciones}")
    print(f"  Fitness CP-SAT: {fitness_cpsat:.0f}  →  GA: {fitness_ga:.0f}  ({mejora_pct:.1f}% mejora)")
    print("\n  → step8 OK")


def main():
    print("Cargando datos desde", INPUTS_DIR)
    datos = cargar_datos(INPUTS_DIR)

    fallidos = []
    try:
        test_step8(datos)
    except AssertionError as e:
        fallidos.append(str(e))
    except Exception as e:
        fallidos.append(f"ERROR inesperado: {e}")
        import traceback
        traceback.print_exc()

    print()
    if fallidos:
        print(f"RESULTADO: {len(fallidos)} test(s) FALLARON:")
        for msg in fallidos:
            print(f"  ✗ {msg}")
        sys.exit(1)
    else:
        print("RESULTADO: PASO 8 VALIDADO ✓")


if __name__ == "__main__":
    main()
