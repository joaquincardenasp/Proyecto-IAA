"""
exporter.py — Exporta el horario generado a Excel (.xlsx).

Hojas:
  "Horario"       — Tabla completa de secciones con sus bloques asignados
  "Plan Común"    — Secciones filtradas por carrera, agrupadas por semestre
  "ICI" … "ICQ"  — Ídem por carrera
  "Métricas"      — Resumen estadístico y fitness de restricciones blandas
  "REPORTE"       — Detalle de todas las violaciones de restricciones (si se pasa reporte)
"""
from __future__ import annotations

import io
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .blocks import TODOS_BLOQUES
from .models import DatosProblema, TipoReunion

# ---------------------------------------------------------------------------
# Paleta y estilos
# ---------------------------------------------------------------------------

_C = {
    "header_dark":  "1F3864",   # azul marino
    "header_light": "2E75B6",   # azul medio
    "CLAS":         "D9E1F2",   # azul claro
    "AYUD":         "E2EFDA",   # verde claro
    "LABT":         "FCE4D6",   # naranja claro
    "sem_banner":   "F2F2F2",   # gris muy claro
    "white":        "FFFFFF",
}

_CARRERAS = ["Plan Común", "ICI", "IOC", "ICE", "ICC", "ICA", "ICQ"]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h: str = "center") -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _bloques_por_dia(bloques_idx: list[int]) -> dict[str, str]:
    """Devuelve {dia: 'HH:MM-HH:MM'} para cada día ocupado."""
    por_dia: dict[str, str] = {}
    for idx in bloques_idx:
        b = TODOS_BLOQUES[idx]
        dia = b.dia.value
        franja = f"{b.hora_inicio}-{b.hora_fin}"
        # Si hay más de un bloque en el mismo día, concatenar
        if dia in por_dia:
            por_dia[dia] = por_dia[dia] + " / " + franja
        else:
            por_dia[dia] = franja
    return por_dia


def _sem_sort_key(s: str) -> tuple:
    """Ordena semestres con posibles sufijos: '1' < '2' < ... < '9a' < '10f'."""
    digits = "".join(c for c in s if c.isdigit())
    return (int(digits) if digits else 999, s)


def _carreras_sems_str(curso) -> tuple[str, str]:
    """Devuelve strings compactos de carreras y semestres del curso."""
    parts_car, parts_sem = [], []
    for car in _CARRERAS:
        sems = curso.semestres_por_carrera.get(car)
        if sems:
            parts_car.append(car)
            parts_sem.append("/".join(sorted(sems, key=_sem_sort_key)))
    return " · ".join(parts_car), " · ".join(parts_sem)


def _header_row(ws, row: int, values: list[str], bg: str,
                font_color: str = "FFFFFF", height: int = 22) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=font_color)
        cell.alignment = _align("center")
        cell.border = _border_thin()
    ws.row_dimensions[row].height = height


def _data_cell(ws, row: int, col: int, value, comp: str = "", align: str = "center") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(_C.get(comp, _C["white"]))
    cell.font = _font()
    cell.alignment = _align(align)
    cell.border = _border_thin()


# ---------------------------------------------------------------------------
# Sheet 1: Horario completo
# ---------------------------------------------------------------------------

def _sheet_horario(wb: openpyxl.Workbook,
                   datos: DatosProblema,
                   asignaciones: dict[str, list[int]]) -> None:
    ws = wb.create_sheet("Horario")
    sec_by_id = {s.id: s for s in datos.secciones}

    HEADERS = ["CODIGO", "TÍTULO", "SECC.", "TIPO", "PROFESOR",
               "LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES",
               "CARRERA(S)", "SEMESTRE(S)"]
    _header_row(ws, 1, HEADERS, _C["header_dark"], height=28)

    sec_ids_ord = sorted(
        asignaciones.keys(),
        key=lambda sid: (
            (sec_by_id[sid].codigo_curso if sid in sec_by_id else ""),
            (sec_by_id[sid].componente.value if sid in sec_by_id else ""),
            (sec_by_id[sid].seccion if sid in sec_by_id else ""),
        ),
    )

    for r, sec_id in enumerate(sec_ids_ord, start=2):
        s = sec_by_id.get(sec_id)
        if not s:
            continue
        curso = datos.cursos.get(s.codigo_curso)
        prof  = datos.profesores.get(s.rut_profesor)
        por_dia = _bloques_por_dia(asignaciones[sec_id])
        cars, sems = _carreras_sems_str(curso) if curso else ("", "")
        comp = s.componente.value

        vals = [
            s.codigo_curso,
            curso.titulo if curso else "",
            s.seccion,
            comp,
            prof.nombre if prof else s.rut_profesor,
            por_dia.get("L", ""),
            por_dia.get("M", ""),
            por_dia.get("X", ""),
            por_dia.get("J", ""),
            por_dia.get("V", ""),
            cars,
            sems,
        ]
        al = ["center","left","center","center","left",
              "center","center","center","center","center","left","center"]
        for col, (val, a) in enumerate(zip(vals, al), start=1):
            _data_cell(ws, r, col, val, comp, a)

    col_widths = [12, 32, 7, 8, 24, 15, 15, 15, 15, 15, 26, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


# ---------------------------------------------------------------------------
# Sheet por carrera
# ---------------------------------------------------------------------------

def _sheet_carrera(wb: openpyxl.Workbook,
                   datos: DatosProblema,
                   asignaciones: dict[str, list[int]],
                   carrera: str) -> None:
    ws = wb.create_sheet(carrera[:31])
    sec_by_id = {s.id: s for s in datos.secciones}

    # Recolectar (semestre, sec_id, Seccion) para esta carrera
    por_sem: dict[str, list[tuple]] = defaultdict(list)
    for sec_id, bloques_idx in asignaciones.items():
        s = sec_by_id.get(sec_id)
        if not s:
            continue
        curso = datos.cursos.get(s.codigo_curso)
        if not curso:
            continue
        sems = curso.semestres_por_carrera.get(carrera, set())
        for sem in sems:
            por_sem[sem].append((sec_id, s))

    # Título del sheet
    n_cols = 11
    title_cell = ws.cell(row=1, column=1, value=f"HORARIO — {carrera.upper()}")
    title_cell.fill = _fill(_C["header_dark"])
    title_cell.font = _font(bold=True, color="FFFFFF", size=12)
    title_cell.alignment = _align("left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 26

    if not por_sem:
        ws.cell(row=2, column=1, value="(sin secciones asignadas)")
        return

    HEADERS = ["SEM.", "CODIGO", "TÍTULO", "SECC.", "TIPO",
               "LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "PROFESOR"]
    _header_row(ws, 2, HEADERS, _C["header_light"], height=20)

    current_row = 3
    for sem in sorted(por_sem.keys(), key=_sem_sort_key):
        # Banner de semestre
        banner = ws.cell(row=current_row, column=1, value=f"  Semestre {sem}")
        banner.fill = _fill(_C["sem_banner"])
        banner.font = _font(bold=True, size=10)
        banner.alignment = _align("left")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=n_cols)
        ws.row_dimensions[current_row].height = 15
        current_row += 1

        seen: set[str] = set()
        secs_sorted = sorted(
            por_sem[sem],
            key=lambda t: (t[1].codigo_curso, t[1].componente.value, t[1].seccion),
        )
        for sec_id, s in secs_sorted:
            if sec_id in seen:
                continue
            seen.add(sec_id)

            curso   = datos.cursos.get(s.codigo_curso)
            prof    = datos.profesores.get(s.rut_profesor)
            por_dia = _bloques_por_dia(asignaciones[sec_id])
            comp    = s.componente.value

            vals = [
                sem, s.codigo_curso, curso.titulo if curso else "",
                s.seccion, comp,
                por_dia.get("L", ""), por_dia.get("M", ""),
                por_dia.get("X", ""), por_dia.get("J", ""),
                por_dia.get("V", ""),
                prof.nombre if prof else s.rut_profesor,
            ]
            al = ["center","center","left","center","center",
                  "center","center","center","center","center","left"]
            for col, (val, a) in enumerate(zip(vals, al), start=1):
                _data_cell(ws, current_row, col, val, comp, a)
            current_row += 1

    col_widths = [6, 12, 30, 7, 7, 15, 15, 15, 15, 15, 24]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet: Métricas
# ---------------------------------------------------------------------------

def _sheet_metricas(wb: openpyxl.Workbook,
                    datos: DatosProblema,
                    asignaciones: dict[str, list[int]],
                    metricas: Optional[dict]) -> None:
    ws = wb.create_sheet("Métricas")
    sec_by_id = {s.id: s for s in datos.secciones}

    def _titulo(row: int, text: str) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = _font(bold=True, color="FFFFFF", size=11)
        cell.fill = _fill(_C["header_dark"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 20

    def _fila(row: int, label: str, value) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True)
        lc.border = _border_thin()
        vc = ws.cell(row=row, column=2, value=value)
        vc.border = _border_thin()

    r = 1
    _titulo(r, "RESUMEN DEL HORARIO GENERADO"); r += 1
    _fila(r, "Secciones asignadas", len(asignaciones)); r += 1
    _fila(r, "Bloques totales",     sum(len(b) for b in asignaciones.values())); r += 1
    _fila(r, "Cursos en el sistema", len(datos.cursos)); r += 1
    _fila(r, "Profesores", len(datos.profesores)); r += 1
    r += 1

    _titulo(r, "DISTRIBUCIÓN POR COMPONENTE"); r += 1
    comp_cnt: dict[str, int] = defaultdict(int)
    for sec_id in asignaciones:
        s = sec_by_id.get(sec_id)
        if s:
            comp_cnt[s.componente.value] += 1
    for comp in ["CLAS", "AYUD", "LABT"]:
        _fila(r, f"  {comp}", comp_cnt.get(comp, 0)); r += 1
    r += 1

    _titulo(r, "DISTRIBUCIÓN POR BLOQUE"); r += 1
    ws.cell(row=r, column=1, value="DÍA").font = _font(bold=True)
    ws.cell(row=r, column=2, value="HORA").font = _font(bold=True)
    ws.cell(row=r, column=3, value="ASIGNACIONES").font = _font(bold=True)
    r += 1
    bloque_cnt: dict[int, int] = defaultdict(int)
    for bloques in asignaciones.values():
        for idx in bloques:
            bloque_cnt[idx] += 1
    for idx in sorted(bloque_cnt):
        b = TODOS_BLOQUES[idx]
        ws.cell(row=r, column=1, value=b.dia.value)
        ws.cell(row=r, column=2, value=f"{b.hora_inicio}-{b.hora_fin}")
        ws.cell(row=r, column=3, value=bloque_cnt[idx])
        r += 1
    r += 1

    if metricas:
        _titulo(r, "FITNESS — RESTRICCIONES BLANDAS"); r += 1
        _fila(r, "Fitness CP-SAT", metricas.get("fitness_cpsat", "N/A")); r += 1
        _fila(r, "Fitness GA",     metricas.get("fitness_ga",    "N/A")); r += 1
        mejora = metricas.get("mejora_pct")
        _fila(r, "Mejora (%)", f"{mejora:.1f}%" if mejora is not None else "N/A"); r += 1
        r += 1
        rb = metricas.get("rb_detalle")
        if rb:
            _titulo(r, "DESGLOSE POR RESTRICCIÓN BLANDA"); r += 1
            for nombre, val in rb.items():
                _fila(r, f"  {nombre}", val); r += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16


# ---------------------------------------------------------------------------
# Sheet: REPORTE de violaciones
# ---------------------------------------------------------------------------

_C_DURA   = "FFDCE1"   # rojo claro
_C_BLANDA = "FFF2CC"   # amarillo claro
_C_DURA_H = "C00000"   # rojo oscuro para header
_C_BLAND_H = "BF8F00"  # amarillo oscuro para header


def _sheet_reporte(wb: openpyxl.Workbook, reporte: dict) -> None:
    ws = wb.create_sheet("REPORTE")

    HEADERS = ["Tipo", "Restricción", "Descripción detallada",
               "Curso / Sección 1", "Curso / Sección 2",
               "Bloque(s)", "Contexto", "Penalización"]

    # Fila de encabezado
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(_C["header_dark"])
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = _align("center")
        cell.border = _border_thin()
    ws.row_dimensions[1].height = 22

    r = 2

    def _escribir_grupo(titulo: str, violaciones: list[dict], bg_header: str) -> None:
        nonlocal r
        if not violaciones:
            return
        # Banner del grupo
        cell = ws.cell(row=r, column=1, value=titulo)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = _fill(bg_header)
        cell.alignment = _align("left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
        ws.row_dimensions[r].height = 16
        r += 1

        for v in violaciones:
            secs = v.get("secciones", [])
            sec1 = f"{secs[0]['codigo']}-{secs[0]['seccion']} {secs[0]['tipo']}" if secs else ""
            sec2 = (
                f"{secs[1]['codigo']}-{secs[1]['seccion']} {secs[1]['tipo']}"
                if len(secs) > 1 else ""
            )
            # Color de fila según si es dura o blanda
            es_dura = not v.get("penalizacion")
            bg = _C_DURA if es_dura else _C_BLANDA

            vals = [
                v["tipo"],
                v["descripcion"],
                v["mensaje"],
                sec1,
                sec2,
                " / ".join(v.get("bloques", [])),
                v.get("contexto", ""),
                v.get("penalizacion") or "",
            ]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = _fill(bg)
                cell.font = _font()
                cell.alignment = _align("left") if col in (3, 4, 5, 7) else _align("center")
                cell.border = _border_thin()
                cell.alignment = Alignment(
                    horizontal="left" if col in (3, 4, 5, 7) else "center",
                    vertical="center",
                    wrap_text=True,
                )
            ws.row_dimensions[r].height = 30
            r += 1

    duras   = reporte.get("violaciones_duras", [])
    blandas = reporte.get("violaciones_blandas", [])
    res     = reporte.get("resumen", {})

    # Resumen en las primeras filas antes de los grupos
    ws.insert_rows(1, amount=5)
    r += 5

    def _rsm(row: int, label: str, value) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _font()

    _rsm(1, "Total violaciones duras (RD1/RD3/RD4)", res.get("total_duras", 0))
    _rsm(2, "Total violaciones blandas (RB1–RB5)",   res.get("total_blandas", 0))
    _rsm(3, "Penalización total blandas",             res.get("penalizacion_total", 0))
    _rsm(4, "Desglose por RB", " | ".join(
        f"{k}: {v:.0f}" for k, v in sorted(res.get("penalizacion_por_rb", {}).items())
    ))
    ws.row_dimensions[5].height = 6   # separador visual

    _escribir_grupo(
        f"RESTRICCIONES DURAS — {res.get('total_duras', 0)} violaciones "
        "(deberían ser 0)",
        duras, _C_DURA_H,
    )
    _escribir_grupo(
        f"RESTRICCIONES BLANDAS — {res.get('total_blandas', 0)} violaciones",
        blandas, _C_BLAND_H,
    )

    col_widths = [8, 28, 60, 28, 28, 28, 32, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A7"  # debajo del resumen y del encabezado de columnas


# ---------------------------------------------------------------------------
# Función pública
# ---------------------------------------------------------------------------

def exportar_horario(
    datos: DatosProblema,
    asignaciones: dict[str, list[int]],
    output_path: str | Path | None = None,
    metricas: Optional[dict] = None,
    reporte: Optional[dict] = None,
) -> bytes:
    """
    Genera un archivo Excel con el horario completo.

    Args:
        datos:        Datos del problema (cursos, secciones, profesores).
        asignaciones: Mapa sec_id → [bloque_idx, ...] (salida del GA o CP-SAT).
        output_path:  Si se da, guarda el .xlsx en esa ruta además de retornar bytes.
        metricas:     Dict opcional: fitness_cpsat, fitness_ga, mejora_pct, rb_detalle.

    Returns:
        Bytes del .xlsx (listos para respuesta HTTP o escritura manual).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)           # eliminar hoja "Sheet" por defecto

    _sheet_horario(wb, datos, asignaciones)
    for carrera in _CARRERAS:
        _sheet_carrera(wb, datos, asignaciones, carrera)
    _sheet_metricas(wb, datos, asignaciones, metricas)
    if reporte:
        _sheet_reporte(wb, reporte)

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    if output_path:
        p = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(excel_bytes)

    return excel_bytes
